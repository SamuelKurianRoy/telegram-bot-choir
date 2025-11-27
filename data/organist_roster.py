# data/organist_roster.py
# Organist Roster Management - Fetch and display organist assignments

import pandas as pd
import io
from datetime import datetime, date, timedelta, timezone
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from data.drive import get_drive_service
from config import get_config
from logging_utils import setup_loggers
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

bot_logger, user_logger = setup_loggers()

# Global cache for organist roster data
_organist_roster_cache = None

def load_organist_roster_data():
    """
    Load organist roster data from Google Sheet and cache it.
    Reads from the 'Order of Songs' sheet.
    
    Returns:
        DataFrame with columns: 'Song/ Responses', 'Name of The Organist'
        or None if failed
    """
    global _organist_roster_cache
    
    try:
        config = get_config()
        roster_sheet_id = config.secrets.get("ORGANIST_ROSTER_SHEET_ID")
        
        if not roster_sheet_id:
            user_logger.error("ORGANIST_ROSTER_SHEET_ID not found in secrets")
            return None
        
        drive_service = get_drive_service()
        
        # Download the Excel file
        request = drive_service.files().export_media(
            fileId=roster_sheet_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_data = io.BytesIO()
        downloader = MediaIoBaseDownload(file_data, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        file_data.seek(0)
        
        # Read the Excel file - specifically the 'Order of Songs' sheet
        df = pd.read_excel(file_data, sheet_name='Order of Songs')
        
        # Ensure required columns exist
        required_columns = ['Song/ Responses', 'Name of The Organist']
        for col in required_columns:
            if col not in df.columns:
                user_logger.error(f"Missing required column: {col}")
                return None
        
        # Cache the data
        _organist_roster_cache = df
        user_logger.info("✅ Organist roster loaded and cached successfully")
        return df
    
    except Exception as e:
        user_logger.error(f"Roster load error: {str(e)[:100]}")
        return None

def get_organist_roster_data():
    """
    Get organist roster data from cache, or load if not cached.
    
    Returns:
        DataFrame with columns: 'Song/ Responses', 'Name of The Organist'
        or None if failed
    """
    global _organist_roster_cache
    
    if _organist_roster_cache is None:
        user_logger.info("Cache miss - loading organist roster from Google Drive")
        return load_organist_roster_data()
    
    user_logger.info(f"Cache hit - using cached organist roster ({len(_organist_roster_cache)} entries)")
    return _organist_roster_cache

def reload_organist_roster():
    """
    Force reload of organist roster data from Google Sheet.
    Called by /refresh command.
    
    Returns:
        bool: True if reload successful, False otherwise
    """
    result = load_organist_roster_data()
    return result is not None

def get_unique_organists():
    """
    Get list of unique organists from the roster.
    
    Returns:
        list: List of unique organist names (excluding empty/NaN values)
    """
    df = get_organist_roster_data()
    
    if df is None:
        return []
    
    try:
        # Get unique organists, excluding NaN values
        organists = df['Name of The Organist'].dropna().unique().tolist()
        # Remove empty strings
        organists = [org for org in organists if str(org).strip()]
        # Sort alphabetically
        organists.sort()
        
        user_logger.info(f"Found {len(organists)} unique organists")
        return organists
    
    except Exception as e:
        user_logger.error(f"Error getting organists: {str(e)[:50]}")
        return []

def get_songs_by_organist(organist_name: str):
    """
    Get all songs assigned to a specific organist.
    
    Args:
        organist_name: Name of the organist
        
    Returns:
        list: List of songs assigned to the organist
    """
    df = get_organist_roster_data()
    
    if df is None:
        return []
    
    try:
        # Filter songs for the specific organist
        organist_songs = df[df['Name of The Organist'] == organist_name]['Song/ Responses'].tolist()
        # Remove NaN values
        organist_songs = [song for song in organist_songs if pd.notna(song)]
        
        user_logger.info(f"Found {len(organist_songs)} songs for {organist_name}")
        return organist_songs
    
    except Exception as e:
        user_logger.error(f"Error getting songs: {str(e)[:50]}")
        return []

def get_unassigned_songs():
    """
    Get all songs that don't have an organist assigned.
    
    Returns:
        list: List of unassigned songs
    """
    df = get_organist_roster_data()
    
    if df is None:
        return []
    
    try:
        # Filter songs where organist is NaN or empty
        unassigned = df[df['Name of The Organist'].isna() | (df['Name of The Organist'].str.strip() == '')]
        unassigned_songs = unassigned['Song/ Responses'].tolist()
        # Remove NaN values from songs
        unassigned_songs = [song for song in unassigned_songs if pd.notna(song)]
        
        user_logger.info(f"Found {len(unassigned_songs)} unassigned songs")
        return unassigned_songs
    
    except Exception as e:
        user_logger.error(f"Error getting unassigned songs: {str(e)[:50]}")
        return []

def get_roster_summary():
    """
    Get summary statistics about the organist roster.
    
    Returns:
        dict: Summary statistics
    """
    df = get_organist_roster_data()
    
    if df is None:
        return {
            'total_songs': 0,
            'assigned_songs': 0,
            'unassigned_songs': 0,
            'total_organists': 0
        }
    
    try:
        total_songs = len(df[df['Song/ Responses'].notna()])
        assigned_songs = len(df[df['Name of The Organist'].notna() & (df['Name of The Organist'].str.strip() != '')])
        unassigned_songs = total_songs - assigned_songs
        total_organists = len(df['Name of The Organist'].dropna().unique())
        
        return {
            'total_songs': total_songs,
            'assigned_songs': assigned_songs,
            'unassigned_songs': unassigned_songs,
            'total_organists': total_organists
        }
    
    except Exception as e:
        user_logger.error(f"Error getting summary: {str(e)[:50]}")
        return {
            'total_songs': 0,
            'assigned_songs': 0,
            'unassigned_songs': 0,
            'total_organists': 0
        }

def get_full_roster_table():
    """
    Get the complete roster table in order.
    
    Returns:
        list: List of tuples (song_name, organist_name) in original order
    """
    df = get_organist_roster_data()
    
    if df is None:
        return []
    
    try:
        # Get all rows with song names (filter out empty songs)
        roster_table = []
        for _, row in df.iterrows():
            song = row['Song/ Responses']
            organist = row['Name of The Organist']
            
            # Only include rows with song names
            if pd.notna(song) and str(song).strip():
                # Handle empty organist
                if pd.isna(organist) or str(organist).strip() == '':
                    organist_name = "Not Assigned"
                else:
                    organist_name = str(organist).strip()
                
                roster_table.append((str(song).strip(), organist_name))
        
        user_logger.info(f"Retrieved full roster table ({len(roster_table)} entries)")
        return roster_table
    
    except Exception as e:
        user_logger.error(f"Error getting full table: {str(e)[:50]}")
        return []

def get_next_sunday():
    """
    Get the next Sunday date. If today is Sunday, return today.
    Uses IST (Indian Standard Time, UTC+5:30) for calculating the current date.
    
    Returns:
        date: The next Sunday date in standard date format (for database lookup)
    """
    # Get current time in IST (UTC+5:30)
    ist_offset = timezone(timedelta(hours=5, minutes=30))
    ist_now = datetime.now(ist_offset)
    today = ist_now.date()
    
    user_logger.info(f"IST now: {ist_now.strftime('%Y-%m-%d %H:%M:%S')}, Today: {today.strftime('%d/%m/%Y')} ({today.strftime('%A')})")
    
    # Monday = 0, Sunday = 6
    if today.weekday() == 6:  # Today is Sunday
        next_sunday = today
        user_logger.info(f"Today is Sunday, using today: {next_sunday.strftime('%d/%m/%Y')}")
    else:
        # Calculate days until next Sunday
        days_ahead = (6 - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        next_sunday = today + timedelta(days=days_ahead)
        user_logger.info(f"Next Sunday is {days_ahead} days away: {next_sunday.strftime('%d/%m/%Y')}")
    
    return next_sunday

def get_songs_for_date(target_date):
    """
    Get songs for a specific date from the main database.
    Uses the same logic as the /date command.
    
    Args:
        target_date: date object
        
    Returns:
        list: List of songs for that date (or next available date)
    """
    try:
        from data.datasets import get_all_data
        
        data = get_all_data()
        df = data["df"]
        
        if df is None or df.empty:
            user_logger.error("Main database is empty")
            return []
        
        user_logger.info(f"Loaded database with {len(df)} rows")
        
        # Ensure 'Date' column is datetime.date
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce').dt.date
        df.dropna(subset=['Date'], inplace=True)
        
        user_logger.info(f"After date parsing: {len(df)} rows")
        
        # Sort dates
        available_dates = sorted(df['Date'].unique())
        user_logger.info(f"Available dates (first 10): {[d.strftime('%d/%m/%Y') for d in available_dates[:10]]}")
        
        # Find songs on the target date
        matching_rows = df[df['Date'] == target_date]
        
        user_logger.info(f"Looking for date: {target_date.strftime('%d/%m/%Y')}, found {len(matching_rows)} matching rows")
        
        if matching_rows.empty:
            # Get the next available date with songs
            next_dates = [d for d in available_dates if d > target_date]
            if not next_dates:
                user_logger.warning(f"No songs found on {target_date.strftime('%d/%m/%Y')} or any later date")
                user_logger.warning(f"Last available date in DB: {available_dates[-1].strftime('%d/%m/%Y') if available_dates else 'None'}")
                return []
            next_date = next_dates[0]
            matching_rows = df[df['Date'] == next_date]
            user_logger.info(f"No songs on {target_date.strftime('%d/%m/%Y')}, using next date: {next_date.strftime('%d/%m/%Y')}")
        
        # Get song columns
        song_columns = [col for col in df.columns if col != 'Date']
        songs = []
        
        # Import IndexFinder to get full song names
        from data.datasets import IndexFinder
        
        for _, row in matching_rows.iterrows():
            for col in song_columns:
                song = row[col]
                if pd.notna(song) and str(song).strip() != '':
                    song_code = str(song).strip()
                    # Get full song name with index
                    song_name = IndexFinder(song_code)
                    full_song = f"{song_code} - {song_name}" if song_name else song_code
                    songs.append(full_song)
        
        user_logger.info(f"Retrieved {len(songs)} songs for date")
        return songs
    
    except Exception as e:
        user_logger.error(f"Error getting songs for date: {str(e)[:100]}")
        return []

def update_songs_for_sunday():
    """
    Update the "Songs for Sunday" sheet with songs from today (if Sunday) or next Sunday.
    Only updates the "Songs" column, leaves "Organist" column unchanged.
    
    Returns:
        tuple: (success: bool, message: str, date_used: date)
    """
    try:
        config = get_config()
        roster_sheet_id = config.secrets.get("ORGANIST_ROSTER_SHEET_ID")
        
        if not roster_sheet_id:
            return False, "ORGANIST_ROSTER_SHEET_ID not found in secrets", None
        
        # Get next Sunday date
        sunday_date = get_next_sunday()
        
        # Get songs for that date
        songs = get_songs_for_date(sunday_date)
        
        user_logger.info(f"Retrieved {len(songs)} songs for {sunday_date.strftime('%d/%m/%Y')}")
        
        if not songs:
            return False, f"No songs found for {sunday_date.strftime('%d/%m/%Y')}", sunday_date
        
        drive_service = get_drive_service()
        
        # Download the current Excel file
        request = drive_service.files().export_media(
            fileId=roster_sheet_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_data = io.BytesIO()
        downloader = MediaIoBaseDownload(file_data, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        file_data.seek(0)
        
        # Load workbook with openpyxl to preserve formulas in other sheets
        wb = load_workbook(file_data)
        
        # Check if "Songs for Sunday" sheet exists
        if "Songs for Sunday" not in wb.sheetnames:
            available_sheets = ', '.join(wb.sheetnames)
            return False, f"'Songs for Sunday' sheet not found. Available: {available_sheets}", sunday_date
        
        user_logger.info(f"Available sheets: {wb.sheetnames}")
        
        # Get the target sheet
        ws = wb["Songs for Sunday"]
        
        # Read current data to preserve organists
        current_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
            current_data.append(row)
        
        user_logger.info(f"Current 'Songs for Sunday' has {len(current_data)} rows")
        
        # Get existing organists (column B, index 1)
        existing_organists = []
        for row in current_data:
            if len(row) > 1:
                existing_organists.append(row[1] if row[1] is not None else '')
            else:
                existing_organists.append('')
        
        # Match the number of organists to the number of songs
        if len(existing_organists) < len(songs):
            # Pad with empty strings
            organists = existing_organists + [''] * (len(songs) - len(existing_organists))
        else:
            # Truncate to match songs
            organists = existing_organists[:len(songs)]
        
        # Clear the sheet data (keep header row)
        ws.delete_rows(2, ws.max_row)
        
        # Write new data
        for i, (song, organist) in enumerate(zip(songs, organists), start=2):
            ws.cell(row=i, column=1, value=song)  # Column A: Songs
            ws.cell(row=i, column=2, value=organist)  # Column B: Organist
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload the file back
        media = MediaIoBaseUpload(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        drive_service.files().update(
            fileId=roster_sheet_id,
            media_body=media
        ).execute()
        
        user_logger.info(f"✅ Updated Songs for Sunday with {len(songs)} songs for {sunday_date.strftime('%d/%m/%Y')}")
        return True, f"✅ Updated {len(songs)} songs for {sunday_date.strftime('%d/%m/%Y')}", sunday_date
    
    except Exception as e:
        user_logger.error(f"Error updating Sunday songs: {str(e)[:100]}")
        return False, f"Error: {str(e)[:100]}", None


def assign_song_to_organist(song_code: str, organist_name: str):
    """
    Assign a specific song to an organist in the Google Sheet.
    
    Args:
        song_code: Song code (e.g., 'H-44', 'L-22')
        organist_name: Name of the organist to assign
    
    Returns:
        tuple: (success: bool, message: str)
    """
    try:
        config = get_config()
        roster_sheet_id = config.secrets.get("ORGANIST_ROSTER_SHEET_ID")
        
        if not roster_sheet_id:
            return False, "ORGANIST_ROSTER_SHEET_ID not found in secrets"
        
        drive_service = get_drive_service()
        
        # Download current file
        request = drive_service.files().export_media(
            fileId=roster_sheet_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_data = io.BytesIO()
        downloader = MediaIoBaseDownload(file_data, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        file_data.seek(0)
        
        # Load workbook using openpyxl to preserve formulas
        wb = load_workbook(file_data)
        
        # Check if 'Order of Songs' sheet exists
        if 'Order of Songs' not in wb.sheetnames:
            return False, "'Order of Songs' sheet not found in workbook"
        
        ws = wb['Order of Songs']
        
        # Find the header row
        header_row = None
        song_col = None
        organist_col = None
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value == 'Song/ Responses':
                    song_col = col_idx
                    header_row = row_idx
                elif cell.value == 'Name of The Organist':
                    organist_col = col_idx
        
        if not song_col or not organist_col:
            return False, "Could not find required columns in sheet"
        
        # Find the song in the sheet
        song_found = False
        song_row = None
        
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=song_col).value
            if cell_value and str(cell_value).strip().upper() == song_code.upper():
                song_row = row_idx
                song_found = True
                break
        
        if not song_found:
            user_logger.warning(f"Song '{song_code}' not found in 'Order of Songs' sheet")
            return False, (
                f"Song '{song_code}' is not in the 'Order of Songs' roster yet.\n\n"
                f"Please add it to the 'Order of Songs' sheet manually first, "
                f"then you can assign an organist to it."
            )
        
        # Update the organist assignment
        ws.cell(row=song_row, column=organist_col, value=organist_name)
        
        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload back to Google Drive
        media = MediaIoBaseUpload(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True
        )
        
        drive_service.files().update(
            fileId=roster_sheet_id,
            media_body=media
        ).execute()
        
        # Invalidate cache to force reload
        global _organist_roster_cache
        _organist_roster_cache = None
        
        user_logger.info(f"✅ Assigned {song_code} to {organist_name}")
        return True, f"✅ {song_code} assigned to {organist_name}"
    
    except Exception as e:
        user_logger.error(f"Error assigning song: {str(e)[:200]}")
        return False, f"Error: {str(e)[:100]}"


def get_songs_for_assignment():
    """
    Get list of songs from 'Songs for Sunday' sheet that can be assigned to organists.
    
    Returns:
        tuple: (success: bool, songs: list of str, message: str)
    """
    try:
        config = get_config()
        roster_sheet_id = config.secrets.get("ORGANIST_ROSTER_SHEET_ID")
        
        if not roster_sheet_id:
            return False, [], "ORGANIST_ROSTER_SHEET_ID not found in secrets"
        
        drive_service = get_drive_service()
        
        # Download the file
        request = drive_service.files().export_media(
            fileId=roster_sheet_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_data = io.BytesIO()
        downloader = MediaIoBaseDownload(file_data, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        file_data.seek(0)
        
        # Read the 'Songs for Sunday' sheet
        try:
            df = pd.read_excel(file_data, sheet_name='Songs for Sunday')
        except Exception as sheet_error:
            return False, [], f"'Songs for Sunday' sheet not found: {str(sheet_error)[:50]}"
        
        # Get songs from the 'Songs' column
        if 'Songs' not in df.columns:
            return False, [], "'Songs' column not found in 'Songs for Sunday' sheet"
        
        # Extract song codes and filter out empty/invalid ones
        songs = []
        for song in df['Songs']:
            if pd.notna(song):
                song_str = str(song).strip()
                if not song_str:
                    continue
                
                # Extract just the song code part (before the " - " separator if present)
                # Format can be: "H-44" or "H-44 - Song Name"
                if ' - ' in song_str:
                    # Has full name, extract just the code
                    song_code = song_str.split(' - ')[0].strip()
                else:
                    # Just the code
                    song_code = song_str
                
                # Validate it's a proper song code (H-XX, L-XX, C-XX)
                if '-' in song_code:
                    parts = song_code.split('-', 1)  # Split only on first dash
                    if len(parts) == 2 and parts[0].upper() in ['H', 'L', 'C'] and parts[1].strip().isdigit():
                        songs.append(song_code.upper())
        
        if not songs:
            # Log what we found for debugging
            sample_songs = [str(s)[:50] for s in df['Songs'].head(5) if pd.notna(s)]
            user_logger.warning(f"No valid song codes found. Sample entries: {sample_songs}")
            return False, [], "No valid song codes found in 'Songs for Sunday' sheet"
        
        return True, songs, f"Found {len(songs)} songs"
    
    except Exception as e:
        user_logger.error(f"Error getting songs for assignment: {str(e)[:200]}")
        return False, [], f"Error: {str(e)[:100]}"
